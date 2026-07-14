# Price provider implementations for fetching asset prices from external APIs.
# Each provider has the same signature: (ticker, start_date?, end_date?) -> PriceResult.
# Providers are stateless — they fetch and return, the service handles storage.
# To swap a provider, change the mapping in asset_price_service._CATEGORY_PROVIDERS.

import asyncio
import logging
from collections.abc import Awaitable, Callable
from datetime import date as date_type
from decimal import Decimal
from typing import NamedTuple

import httpx

logger = logging.getLogger(__name__)

# --- Result types ---

PriceResult = list[tuple[date_type, Decimal, str]]

# CEDEAR ratio result: (ticker, underlying, ratio).
RatioResult = list[tuple[str, str, Decimal]]


# CEDEAR ratio fetch result: ratios + source date (if parseable).
class RatioFetchResult(NamedTuple):
    ratios: RatioResult
    source_date: date_type | None


# --- Provider metadata ---


# Describes a price provider: its source name, fetch function, and capabilities.
class PriceProviderInfo(NamedTuple):
    source: str
    fetch: Callable[..., Awaitable[PriceResult]]
    supports_history: bool


# --- Source name constants (stored in the source column of asset_prices/cedear_ratios) ---

SOURCE_YFINANCE = "yfinance"
SOURCE_COINGECKO = "coingecko"
SOURCE_CAFCI = "cafci"
COMAFI_SOURCE = "comafi"
BYMA_SOURCE = "byma"


# --- Price providers ---


# Fetches prices from Yahoo Finance via yfinance for stocks, CEDEARs, and government bonds.
# Returns daily closing prices for the given period.
async def fetch_yfinance(
    ticker: str,
    start_date: date_type | None = None,
    end_date: date_type | None = None,
) -> PriceResult:
    import asyncio

    import yfinance as yf

    def _fetch() -> PriceResult:
        from datetime import timedelta

        t = yf.Ticker(ticker)
        kwargs: dict = {}
        if start_date and end_date:
            kwargs["start"] = start_date.isoformat()
            # yfinance end date is exclusive — add one day to include the target date.
            kwargs["end"] = (end_date + timedelta(days=1)).isoformat()
        else:
            kwargs["period"] = "5d"
        hist = t.history(**kwargs)
        if hist.empty:
            return []
        # Determine currency from ticker info (fallback to USD).
        try:
            currency = t.info.get("currency", "USD").upper()
        except Exception:
            currency = "USD"
        results: PriceResult = []
        for idx, row in hist.iterrows():
            price_date = idx.date() if hasattr(idx, "date") else idx
            close = row.get("Close")
            if close is not None:
                results.append((price_date, Decimal(str(round(close, 6))), currency))
        return results

    try:
        return await asyncio.to_thread(_fetch)
    except Exception:
        logger.exception("yfinance fetch failed for %s.", ticker)
        return []


# Fetches prices from CoinGecko for crypto assets.
# ticker should be a CoinGecko coin id (e.g. "bitcoin", "ethereum").
# start_date/end_date are accepted for signature uniformity but ignored — CoinGecko
# always returns the last 7 days via the market_chart endpoint.
async def fetch_coingecko(
    ticker: str,
    start_date: date_type | None = None,
    end_date: date_type | None = None,
) -> PriceResult:
    url = f"https://api.coingecko.com/api/v3/coins/{ticker}/market_chart"
    params = {"vs_currency": "usd", "days": "7", "interval": "daily"}
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            response = await client.get(url, params=params)
            response.raise_for_status()
            data = response.json()
    except httpx.HTTPError:
        logger.exception("CoinGecko fetch failed for %s.", ticker)
        return []

    prices = data.get("prices", [])
    results: PriceResult = []
    for timestamp_ms, price in prices:
        price_date = date_type.fromtimestamp(timestamp_ms / 1000)
        results.append((price_date, Decimal(str(round(price, 6))), "USD"))
    return results


# --- CEDEAR ratio provider ---

# Banco Comafi Excel configuration.
COMAFI_CEDEAR_URL = "https://www.comafi.com.ar/Multimedios/otros/7279.xlsx"
COMAFI_TIMEOUT = 30.0
COMAFI_HEADER_SCAN_MAX_ROW = 15
COMAFI_TICKER_HEADER_KEYWORD = "mercado"
COMAFI_TICKER_HEADER_KEYWORD_2 = "identif"
COMAFI_RATIO_HEADER_KEYWORD = "ratio"
COMAFI_BYMA_SUFFIX = ".BA"
COMAFI_RATIO_SEPARATOR = ":"


# Fetches all CEDEAR ratios from Banco Comafi's Excel file.
# Returns ratios + the internal date from the spreadsheet (for freshness comparison).
async def fetch_comafi_ratios() -> RatioFetchResult:
    import asyncio
    import io

    try:
        async with httpx.AsyncClient(timeout=COMAFI_TIMEOUT) as client:
            response = await client.get(COMAFI_CEDEAR_URL)
            response.raise_for_status()
            content = response.content
    except httpx.HTTPError:
        logger.exception("Comafi CEDEAR Excel fetch failed.")
        return RatioFetchResult([], None)

    def _parse(data: bytes) -> RatioFetchResult:
        import re

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return RatioFetchResult([], None)

        # Parse the internal date from the first rows (e.g., "LISTA TOTAL DE CEDEARS AL  13.03.2025").
        source_date: date_type | None = None
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if isinstance(cell, date_type):
                    source_date = cell
                    break
                if isinstance(cell, str):
                    match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", cell)
                    if match:
                        try:
                            source_date = date_type(int(match.group(3)), int(match.group(2)), int(match.group(1)))
                        except ValueError:
                            pass
            if source_date:
                break

        # Find the header row to locate columns dynamically.
        header_row = None
        ticker_col = None
        ratio_col = None
        for row in ws.iter_rows(min_row=1, max_row=COMAFI_HEADER_SCAN_MAX_ROW, values_only=False):
            for cell in row:
                val = str(cell.value or "").strip().lower()
                if COMAFI_TICKER_HEADER_KEYWORD in val and COMAFI_TICKER_HEADER_KEYWORD_2 in val:
                    header_row = cell.row
                    ticker_col = cell.column
                if COMAFI_RATIO_HEADER_KEYWORD in val:
                    ratio_col = cell.column
            if header_row and ticker_col and ratio_col:
                break

        if not header_row or not ticker_col or not ratio_col:
            logger.warning("Could not find header columns in Comafi Excel.")
            wb.close()
            return RatioFetchResult([], source_date)

        results: RatioResult = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            ticker_cell = row[ticker_col - 1].value if len(row) >= ticker_col else None
            ratio_cell = row[ratio_col - 1].value if len(row) >= ratio_col else None
            if not ticker_cell or not ratio_cell:
                continue

            ticker_str = str(ticker_cell).strip().upper()
            if not ticker_str:
                continue

            # Parse ratio — formats: "10:1", "10", "10.0".
            ratio_str = str(ratio_cell).strip().replace(",", ".")
            if COMAFI_RATIO_SEPARATOR in ratio_str:
                parts = ratio_str.split(COMAFI_RATIO_SEPARATOR)
                try:
                    ratio_val = Decimal(parts[0].strip()) / Decimal(parts[1].strip())
                except Exception:
                    continue
            else:
                try:
                    ratio_val = Decimal(ratio_str)
                except Exception:
                    continue

            if ratio_val <= 0:
                continue

            # Build BYMA ticker (add .BA suffix if not present).
            cedear_ticker = ticker_str if ticker_str.endswith(COMAFI_BYMA_SUFFIX) else f"{ticker_str}{COMAFI_BYMA_SUFFIX}"
            # Underlying is the ticker without .BA.
            underlying = ticker_str.replace(COMAFI_BYMA_SUFFIX, "")

            results.append((cedear_ticker, underlying, ratio_val))

        wb.close()
        return RatioFetchResult(results, source_date)

    try:
        return await asyncio.to_thread(_parse, content)
    except Exception:
        logger.exception("Failed to parse Comafi CEDEAR Excel.")
        return RatioFetchResult([], None)


# --- BYMA PDF ratio provider ---

BYMA_CEDEARS_PAGE_URL = "https://www.byma.com.ar/productos/productos-financieros/cedears"
BYMA_PDF_CDN_PREFIX = "https://cdn.prod.website-files.com/"
BYMA_TIMEOUT = 30.0
BYMA_RATIO_SEPARATOR = ":"
BYMA_BYMA_SUFFIX = ".BA"


# Fetches all CEDEAR ratios from the BYMA PDF.
# First discovers the current PDF URL from the BYMA page, then parses the PDF.
# Returns ratios + the date from the PDF filename (for freshness comparison).
async def fetch_byma_ratios() -> RatioFetchResult:
    import asyncio

    # Step 1: Discover the PDF URL from the BYMA page.
    try:
        async with httpx.AsyncClient(timeout=BYMA_TIMEOUT, follow_redirects=True) as client:
            page_response = await client.get(BYMA_CEDEARS_PAGE_URL)
            page_response.raise_for_status()
            page_html = page_response.text
    except httpx.HTTPError:
        logger.exception("BYMA CEDEARs page fetch failed.")
        return RatioFetchResult([], None)

    # Find the PDF URL in the page HTML (CDN link with "CEDEARs" in the filename).
    import re

    pdf_match = re.search(
        r'(https://cdn\.prod\.website-files\.com/[^"\']+CEDEARs[^"\']*\.pdf)',
        page_html,
        re.IGNORECASE,
    )
    if not pdf_match:
        logger.warning("Could not find CEDEAR PDF link on BYMA page.")
        return RatioFetchResult([], None)

    pdf_url = pdf_match.group(1)

    # Parse date from the PDF filename (e.g., "BYMA-CEDEARs-2026-02-03.pdf").
    source_date: date_type | None = None
    date_match = re.search(r"(\d{4})-(\d{2})-(\d{2})\.pdf$", pdf_url)
    if date_match:
        try:
            source_date = date_type(int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3)))
        except ValueError:
            pass

    # Step 2: Download the PDF.
    try:
        async with httpx.AsyncClient(timeout=BYMA_TIMEOUT) as client:
            pdf_response = await client.get(pdf_url)
            pdf_response.raise_for_status()
            pdf_content = pdf_response.content
    except httpx.HTTPError:
        logger.exception("BYMA CEDEAR PDF download failed: %s", pdf_url)
        return RatioFetchResult([], source_date)

    # Step 3: Parse the PDF.
    def _parse(data: bytes) -> RatioResult:
        import io

        import pdfplumber

        results: RatioResult = []
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                for line in text.split("\n"):
                    line = line.strip()
                    if not line or "Ratio" in line or "Nombre" in line or "Bolsas y Mercados" in line:
                        continue

                    # Format: "Company Name  TICKER  EXCHANGE  RATIO:1"
                    # The ratio is always at the end, in "N:N" format.
                    ratio_match = re.search(r"(\d+):(\d+)\s*$", line)
                    if not ratio_match:
                        continue

                    # Extract ticker — it's the uppercase word before the exchange name.
                    # Split line into parts and find the ticker.
                    before_ratio = line[: ratio_match.start()].strip()
                    parts = before_ratio.split()
                    if len(parts) < 2:
                        continue

                    # Ticker is typically the second-to-last or third-to-last token before ratio.
                    # The exchange (NYSE, NASDAQ, etc.) is right before the ratio.
                    # Walk backwards: last part = exchange, second-to-last = ticker.
                    ticker = None
                    for i in range(len(parts) - 1, -1, -1):
                        token = parts[i].strip()
                        # Skip exchange names and partial exchange names.
                        if token.upper() in {
                            "NYSE",
                            "NASDAQ",
                            "XETRA",
                            "FRANKFURT",
                            "B3",
                            "ARCA",
                            "GS",
                            "GM",
                            "AMERICAN",
                        }:
                            continue
                        # Found the ticker.
                        ticker = token.upper()
                        break

                    if not ticker or not any(c.isalnum() for c in ticker):
                        continue

                    # Parse ratio.
                    numerator = int(ratio_match.group(1))
                    denominator = int(ratio_match.group(2))
                    if denominator == 0:
                        continue
                    ratio_val = Decimal(numerator) / Decimal(denominator)

                    if ratio_val <= 0:
                        continue

                    cedear_ticker = ticker if ticker.endswith(BYMA_BYMA_SUFFIX) else f"{ticker}{BYMA_BYMA_SUFFIX}"
                    underlying = ticker.replace(BYMA_BYMA_SUFFIX, "")

                    results.append((cedear_ticker, underlying, ratio_val))

        return results

    try:
        ratios = await asyncio.to_thread(_parse, pdf_content)
        return RatioFetchResult(ratios, source_date)
    except Exception:
        logger.exception("Failed to parse BYMA CEDEAR PDF.")
        return RatioFetchResult([], source_date)


# --- FCI (mutual fund) price provider ---

CAFCI_EXCEL_URL = "https://api.pub.cafci.org.ar/pb_get"
CAFCI_TIMEOUT = 30.0
CAFCI_HEADER_SCAN_MAX_ROW = 15
CAFCI_CODE_HEADER = "código cafci"
CAFCI_VCP_HEADER = "valor"
CAFCI_DATE_HEADER = "fecha"
CAFCI_CURRENCY_HEADER = "moneda fondo"

SOURCE_ARGENTIADATOS = "argentinadatos"
ARGENTIADATOS_BASE = "https://api.argentinadatos.com/v1/finanzas/fci"
ARGENTIADATOS_TYPES = ["mercadoDinero", "rentaFija", "rentaVariable", "rentaMixta", "otros"]
ARGENTIADATOS_TIMEOUT = 15.0

# Module-level caches. Populated on first fetch per process lifecycle.
# _cafci_prices: code → (date, price, currency). All funds from the latest Excel download.
# _cafci_registry: code → fund name. Used for ArgentinaDatos fallback.
_cafci_prices: dict[str, tuple[date_type, Decimal, str]] | None = None
_cafci_registry: dict[str, str] | None = None
_cafci_lock: asyncio.Lock | None = None


# Returns the process-wide CAFCI download lock, creating it lazily on first use.
def _get_cafci_lock() -> asyncio.Lock:
    global _cafci_lock
    if _cafci_lock is None:
        _cafci_lock = asyncio.Lock()
    return _cafci_lock


# Clears the CAFCI cache. Called before each refresh cycle so the next fetch re-downloads.
def clear_fci_cache() -> None:
    global _cafci_prices, _cafci_registry
    _cafci_prices = None
    _cafci_registry = None


# Fetches the latest FCI cuotaparte price for a given CAFCI code.
# Primary: CAFCI public Excel (cached per refresh cycle). Fallback: ArgentinaDatos JSON API.
async def fetch_fci(
    ticker: str,
    start_date: date_type | None = None,
    end_date: date_type | None = None,
) -> PriceResult:
    # Try cached CAFCI data first (populated by the first call in the refresh cycle).
    if _cafci_prices is not None:
        entry = _cafci_prices.get(ticker)
        return [entry] if entry else []

    # First call — download and cache. Lock ensures only one download even with concurrent calls.
    async with _get_cafci_lock():
        # Re-check after acquiring lock (another coroutine may have populated the cache).
        if _cafci_prices is not None:
            entry = _cafci_prices.get(ticker)
            return [entry] if entry else []
        await _load_cafci_cache()

    if _cafci_prices is not None:
        entry = _cafci_prices.get(ticker)
        if entry:
            return [entry]

    logger.warning("CAFCI Excel returned no price for ticker %s. Trying ArgentinaDatos.", ticker)
    return await _fetch_fci_from_argentinadatos(ticker)


# Downloads the CAFCI Excel and caches all fund prices + registry in module-level dicts.
# Called once per refresh cycle — subsequent fetch_fci() calls read from cache.
async def _load_cafci_cache() -> None:
    global _cafci_prices, _cafci_registry
    import asyncio
    import io

    try:
        async with httpx.AsyncClient(timeout=CAFCI_TIMEOUT) as client:
            response = await client.get(CAFCI_EXCEL_URL)
            response.raise_for_status()
            content = response.content
    except httpx.HTTPError:
        logger.exception("CAFCI Excel fetch failed.")
        _cafci_prices = {}
        return

    def _parse(data: bytes) -> tuple[dict[str, tuple[date_type, Decimal, str]], dict[str, str]]:
        import re

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return {}, {}

        # Find the header row by locating the CAFCI code column, then map all columns from that row.
        code_col = None
        vcp_col = None
        date_col = None
        currency_col = None
        header_row = None

        for row in ws.iter_rows(min_row=1, max_row=CAFCI_HEADER_SCAN_MAX_ROW, values_only=False):
            for cell in row:
                if CAFCI_CODE_HEADER in str(cell.value or "").strip().lower():
                    header_row = cell.row
                    break
            if header_row:
                for cell in row:
                    val = str(cell.value or "").strip().lower()
                    if CAFCI_CODE_HEADER in val:
                        code_col = cell.column
                    if CAFCI_VCP_HEADER in val:
                        vcp_col = cell.column
                    if CAFCI_DATE_HEADER in val:
                        date_col = cell.column
                    if CAFCI_CURRENCY_HEADER in val:
                        currency_col = cell.column
                break

        if not header_row or not code_col or not vcp_col:
            logger.warning("Could not find header columns in CAFCI Excel.")
            wb.close()
            return {}, {}

        # Parse all rows into the prices dict and registry.
        prices: dict[str, tuple[date_type, Decimal, str]] = {}
        registry: dict[str, str] = {}

        for row in ws.iter_rows(min_row=header_row + 2, values_only=False):
            code_cell = row[code_col - 1].value if len(row) >= code_col else None
            if not code_cell:
                continue

            code_str = str(int(code_cell)) if isinstance(code_cell, (int, float)) else str(code_cell).strip()
            fund_name = str(row[0].value or "").strip()
            if fund_name:
                registry[code_str] = fund_name

            # Extract VCP.
            vcp_cell = row[vcp_col - 1].value if len(row) >= vcp_col else None
            if not vcp_cell:
                continue
            try:
                vcp = Decimal(str(vcp_cell).strip().replace(",", "."))
            except Exception:
                continue
            if vcp <= 0:
                continue

            # Extract date.
            price_date = None
            if date_col:
                date_cell = row[date_col - 1].value if len(row) >= date_col else None
                if isinstance(date_cell, date_type):
                    price_date = date_cell
                elif date_cell:
                    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", str(date_cell))
                    if match:
                        y = int(match.group(3))
                        if y < 100:
                            y += 2000
                        try:
                            price_date = date_type(y, int(match.group(2)), int(match.group(1)))
                        except ValueError:
                            pass

            if not price_date:
                price_date = date_type.today()

            # Extract currency.
            currency = "ARS"
            if currency_col:
                curr_cell = row[currency_col - 1].value if len(row) >= currency_col else None
                if curr_cell:
                    currency = str(curr_cell).strip().upper()

            prices[code_str] = (price_date, vcp, currency)

        wb.close()
        return prices, registry

    try:
        prices, registry = await asyncio.to_thread(_parse, content)
        _cafci_prices = prices
        _cafci_registry = registry
        logger.info("CAFCI cache loaded: %d prices, %d registry entries.", len(prices), len(registry))
    except Exception:
        logger.exception("Failed to parse CAFCI Excel.")
        _cafci_prices = {}


# Fetches FCI price from the ArgentinaDatos JSON API (fallback).
# Requires _cafci_registry to map CAFCI code → fund name.
async def _fetch_fci_from_argentinadatos(ticker: str) -> PriceResult:
    global _cafci_registry

    # Need the fund name to search in ArgentinaDatos.
    fund_name = (_cafci_registry or {}).get(ticker)
    if not fund_name:
        logger.warning("No fund name in registry for CAFCI code %s. Cannot query ArgentinaDatos.", ticker)
        return []

    # Fetch all types in parallel.
    async def _fetch_type(fci_type: str) -> list[dict]:
        try:
            async with httpx.AsyncClient(timeout=ARGENTIADATOS_TIMEOUT) as client:
                response = await client.get(f"{ARGENTIADATOS_BASE}/{fci_type}/ultimo")
                response.raise_for_status()
                return response.json()
        except Exception:
            return []

    import asyncio

    all_results = await asyncio.gather(*[_fetch_type(t) for t in ARGENTIADATOS_TYPES])

    # Search all responses for the fund name.
    fund_name_lower = fund_name.lower()
    for items in all_results:
        for item in items:
            fondo = (item.get("fondo") or "").strip()
            if fondo.lower() == fund_name_lower:
                vcp = item.get("vcp")
                fecha = item.get("fecha")
                if not vcp or not fecha:
                    continue
                try:
                    price_date = date_type.fromisoformat(fecha)
                    price_val = Decimal(str(vcp))
                    return [(price_date, price_val, "ARS")]
                except Exception:
                    continue

    logger.warning("Fund '%s' not found in ArgentinaDatos.", fund_name)
    return []
