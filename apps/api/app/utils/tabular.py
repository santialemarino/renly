# Tabular file parsing (CSV / XLSX) into a header row + string data rows. Generic; used by imports.

import csv
import io
from datetime import date as date_type
from datetime import datetime

from openpyxl import load_workbook

# Bytes scanned to sniff the CSV delimiter (handles comma, semicolon, and tab separated files).
_SNIFF_SAMPLE_BYTES = 4096
_CSV_DELIMITERS = ",;\t"


# Renders a cell value as a trimmed string the import layer can validate (dates ISO, ints without .0).
def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        # A midnight datetime is a date that was exported with a time component; show it as a date.
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date_type):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


# Reads CSV bytes into a raw grid. Tolerates a UTF-8 BOM and falls back to latin-1; sniffs delimiter.
def _read_csv(content: bytes) -> list[list[str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    try:
        dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(text[:_SNIFF_SAMPLE_BYTES], delimiters=_CSV_DELIMITERS)
    except csv.Error:
        dialect = csv.get_dialect("excel")
    return [list(row) for row in csv.reader(io.StringIO(text), dialect)]


# Reads the active sheet of an XLSX file into a raw grid.
def _read_xlsx(content: bytes) -> list[list[object]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        workbook.close()
        raise ValueError("The spreadsheet has no readable sheet.")
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


# Normalizes a raw grid into (headers, rows): drops blank rows, trims headers, pads rows to width.
def _finalize(grid: list[list[object]]) -> tuple[list[str], list[list[str]]]:
    stringified = [[_cell_to_str(cell) for cell in row] for row in grid]
    non_empty = [row for row in stringified if any(cell != "" for cell in row)]
    if not non_empty:
        raise ValueError("The file is empty.")
    headers = [header.strip() for header in non_empty[0]]
    width = len(headers)
    rows = [(row + [""] * width)[:width] for row in non_empty[1:]]
    return headers, rows


# Parses a CSV/TSV or XLSX upload into (headers, rows). Raises ValueError on unsupported/unreadable
# files. CSV and TSV share the reader — the delimiter (comma/semicolon/tab) is sniffed, not assumed.
def parse_tabular(filename: str, content: bytes) -> tuple[list[str], list[list[str]]]:
    name = filename.lower()
    if name.endswith((".csv", ".tsv")):
        return _finalize(_read_csv(content))
    if name.endswith(".xlsx"):
        return _finalize(_read_xlsx(content))
    raise ValueError("Unsupported file type. Upload a .csv, .tsv, or .xlsx file.")
